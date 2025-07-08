# --- Imports nécessaires pour l'ensemble du script ---
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
import hashlib
import collections
import re
import zipfile
import shutil
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
from datetime import datetime

# --- Configuration du répertoire de sortie commun ---
output_directory = r"C:\Users\PREDATOR_PC\OneDrive\Bureau\Data Extractor"

# Créer le répertoire de sortie s'il n'existe pas
os.makedirs(output_directory, exist_ok=True)
print(f"Répertoire de sortie configuré : {output_directory}")

# --- Fonctions Helper communes ---

def get_distinct_color(identifier):
    """
    Génère une couleur de fond distincte basée sur un identifiant.
    Utilise un hachage pour une distribution raisonnable des couleurs pâles.
    Retourne un code hexadécimal 6 chiffres (sans alpha).
    """
    if not identifier:
        return "FFFFFF"  # Blanc si pas d'identifiant

    hash_object = hashlib.sha256(str(identifier).encode())
    hex_dig = hash_object.hexdigest()

    r = int(hex_dig[:2], 16)
    g = int(hex_dig[2:4], 16)
    b = int(hex_dig[4:6], 16)

    # Ajuster pour couleurs pâles (200–255)
    base_color_value = 200
    range_size = 56  # 255 - 200 + 1

    r = min(255, base_color_value + (r % range_size))
    g = min(255, base_color_value + (g % range_size))
    b = min(255, base_color_value + (b % range_size))

    return f"{r:02X}{g:02X}{b:02X}"

def read_file_with_multiple_encodings(file_path, encodings=['utf-16', 'utf-16-le', 'utf-8-sig', 'utf-8']):
    """Tente de lire un fichier en utilisant une liste d'encodages donnés."""
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
                return content
        except (UnicodeDecodeError, Exception):
            continue
    print(f"Erreur : Impossible de lire le fichier {os.path.basename(file_path)} avec les encodages essayés.")
    return None

def normalize_expression(expression):
    """Convertit une expression (liste ou chaîne multiligne) en une seule chaîne lisible."""
    if isinstance(expression, list):
        return "\n".join(str(item).strip() for item in expression if item)
    elif isinstance(expression, str):
        return expression.strip()
    return "N/A"

# --- Fonctions pour l'extraction des tables et colonnes ---

def extract_table_column_names(json_data):
    """
    Extrait le nom de la table et le nom de chaque colonne du schéma JSON,
    en excluant les tables où isHidden est True.
    Retourne un DataFrame.
    """
    records = []

    model_info = json_data.get("model", {})

    if "tables" in model_info:
        for table in model_info["tables"]:
            if table.get("isHidden", False):
                continue

            table_name = table.get("name", "Table sans nom")

            if "columns" in table:
                for column in table["columns"]:
                    col_name = column.get("name", "Colonne sans nom")

                    records.append({
                        "Nom de la Table": table_name,
                        "Nom de la colonne": col_name
                    })

    return pd.DataFrame(records)

def run_tables_columns_extraction(datamodelschema_json_path, output_directory):
    """Exécute l'extraction des tables/colonnes visibles et retourne le DataFrame."""
    print(f"\n{'='*50}")
    print("Début de l'extraction des tables et colonnes visibles à partir du fichier JSON.")
    print(f"Source JSON : {datamodelschema_json_path}")

    if not datamodelschema_json_path or not os.path.exists(datamodelschema_json_path):
        print("Erreur : Le fichier DataModelSchema.json n'a pas été trouvé pour l'extraction des tables et colonnes.")
        return None

    try:
        with open(datamodelschema_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        df = extract_table_column_names(data)

        if df.empty:
            df = pd.DataFrame(columns=["Nom de la Table", "Nom de la colonne"])
            print("Aucune table ou colonne visible trouvée dans le fichier JSON.")
        else:
            print(f"{len(df)} lignes (colonnes) trouvées pour les tables visibles.")

        return df

    except FileNotFoundError:
        print(f"Erreur : Le fichier spécifié est introuvable à {datamodelschema_json_path}.")
        return None
    except json.JSONDecodeError:
        print(f"Erreur : Impossible de décoder le fichier JSON. Assurez-vous qu'il est correctement formaté.")
        return None
    except ImportError:
        print("Erreur : Les bibliothèques nécessaires (pandas) ne sont pas installées.")
        print("Veuillez les installer en exécutant : pip install pandas")
        return None
    except Exception as e:
        print(f"Une erreur est survenue lors de l'extraction des tables et colonnes : {e}")
        return None

# --- Fonctions pour l'extraction des données structurées ---

def extract_source_info_from_m_expression(expression):
    """
    Tente d'extraire la source (chemin fichier/connexion) et le nom de la table source
    d'une expression M.
    """
    source_path = "N/A"
    source_table_name = "N/A"
    filter_steps = "N/A"
    expression_str = "N/A"

    if isinstance(expression, list):
        expression_str = "\n".join(expression)
    elif isinstance(expression, str):
        expression_str = expression
    else:
        return source_path, source_table_name, expression_str, filter_steps

    display_expression = expression_str
    if len(display_expression) > 500:
        display_expression = display_expression[:500] + "..."

    file_match = re.search(r'File\.Contents\("([^"]+)"\)', expression_str)
    if file_match:
        source_path = file_match.group(1)

    item_match = re.search(r'Source\{\[Item="([^"]+)",Kind="Table"\]\}', expression_str)
    if item_match:
        source_table_name = item_match.group(1)
    elif source_path != "N/A":
        try:
            base_name = os.path.basename(source_path)
            source_table_name = os.path.splitext(base_name)[0]
        except:
            pass

    filter_matches = re.findall(r'Table\.SelectRows\(.*?, (.*?)\)', expression_str)
    if filter_matches:
        filter_steps = " ; ".join(filter_matches)

    return source_path, source_table_name, display_expression, filter_steps

def process_data_model_for_structured_sheet(json_data):
    """
    Extrait les données du modèle Power BI et les organise par type d'entité
    pour la génération du rapport structuré en une seule feuille.
    Retourne un dictionnaire de DataFrames.
    """
    model_info = json_data.get("model", {})

    tables_list = []
    columns_list = []
    measures_list = []
    partitions_list = []
    relationships_list = []
    cultures_list = []
    hierarchies_list = []
    hierarchy_levels_list = []
    annotations_list = []
    variations_list = []

    if "annotations" in model_info:
        for annotation in model_info["annotations"]:
            annotations_list.append({
                "Type Entité Parente": "Modèle",
                "Nom Entité Parente": json_data.get("name", "Modèle sans nom"),
                "Nom Annotation": annotation.get("name", "N/A"),
                "Valeur Annotation": str(annotation.get("value", "N/A")),
                "Nom Tableau Parent": "N/A",
                "Nom Colonne Parent": "N/A",
                "Nom Mesure Parent": "N/A",
                "Nom Relation Parente": "N/A",
                "Nom Hiérarchie Parente": "N/A",
                "Nom Niveau Hiérarchie Parente": "N/A",
                "Nom Partition Parente": "N/A"
            })

    if "tables" in model_info:
        for table in model_info["tables"]:
            table_name = table.get("name", "Table sans nom")

            tables_list.append({
                "Nom Tableau": table_name,
                "isHidden": table.get("isHidden", False),
                "isPrivate": table.get("isPrivate", False),
                "showAsVariationsOnly": table.get("showAsVariationsOnly", False),
                "lineageTag": table.get("lineageTag", "N/A"),
                "description": table.get("description", "N/A"),
                "Partitions (Noms)": ", ".join([p.get("name", "Partition sans nom") for p in table.get("partitions", [])]),
                "Colonnes (Noms)": ", ".join([c.get("name", "Colonne sans nom") for c in table.get("columns", [])]),
                "Mesures (Noms)": ", ".join([m.get("name", "Mesure sans nom") for m in table.get("measures", [])]),
                "Hiérarchies (Noms)": ", ".join([h.get("name", "Hiérarchie sans nom") for h in table.get("hierarchies", [])]),
            })

            if "partitions" in table:
                for i, partition in enumerate(table["partitions"]):
                    partition_name = partition.get("name", f"Partition {i+1}")
                    partition_mode = partition.get("mode", "N/A")
                    partition_source_type = "N/A"
                    partition_expression_display = "N/A"
                    part_source_data = "N/A"
                    part_source_table_name = "N/A"
                    part_filter_steps = "N/A"

                    if "source" in partition:
                        source = partition["source"]
                        partition_source_type = source.get("type", "N/A")
                        expression = source.get("expression")
                        if expression:
                            part_source_data, part_source_table_name, partition_expression_display, part_filter_steps = extract_source_info_from_m_expression(expression)
                            partition_expression_display = normalize_expression(partition_expression_display)

                    partitions_list.append({
                        "Nom Partition": partition_name,
                        "Nom Tableau Parent": table_name,
                        "mode": partition_mode,
                        "source.type": partition_source_type,
                        "source.expression (Tronqué)": partition_expression_display,
                        "Source de Données (Extrait)": part_source_data,
                        "Nom Table Source (Extrait)": part_source_table_name,
                        "Filtres (Extrait)": part_filter_steps,
                        "lineageTag": partition.get("lineageTag", "N/A")
                    })

                    if "annotations" in partition:
                        for annotation in partition["annotations"]:
                            annotations_list.append({
                                "Type Entité Parente": "Partition",
                                "Nom Entité Parente": partition_name,
                                "Nom Annotation": annotation.get("name", "N/A"),
                                "Valeur Annotation": str(annotation.get("value", "N/A")),
                                "Nom Tableau Parent": table_name,
                                "Nom Colonne Parent": "N/A",
                                "Nom Mesure Parent": "N/A",
                                "Nom Relation Parente": "N/A",
                                "Nom Hiérarchie Parente": "N/A",
                                "Nom Niveau Hiérarchie Parente": "N/A",
                                "Nom Partition Parente": partition_name
                            })

            if "columns" in table:
                for column in table["columns"]:
                    col_name = column.get("name", "Colonne sans nom")
                    columns_list.append({
                        "Nom Colonne": col_name,
                        "Nom Tableau Parent": table_name,
                        "dataType": column.get("dataType", "N/A"),
                        "sourceColumn": column.get("sourceColumn", "N/A"),
                        "summarizeBy": column.get("summarizeBy", "N/A"),
                        "isHidden": column.get("isHidden", False),
                        "isNameInferred": column.get("isNameInferred", False),
                        "dataCategory": column.get("dataCategory", "N/A"),
                        "formatString": column.get("formatString", "N/A"),
                        "sortByColumn": column.get("sortByColumn", "N/A"),
                        "lineageTag": column.get("lineageTag", "N/A"),
                        "expression": normalize_expression(column.get("expression", "N/A")),
                        "description": column.get("description", "N/A"),
                        "Annotations (Noms)": ", ".join([a.get("name", "Annotation sans nom") for a in column.get("annotations", [])]),
                        "Variations (Noms)": ", ".join([v.get("name", "Variation sans nom") for v in column.get("variations", [])]),
                    })

                    if "variations" in column:
                        for variation in column["variations"]:
                            variation_name = variation.get("name", "Variation sans nom")
                            variations_list.append({
                                "Nom Variation": variation_name,
                                "Nom Colonne Parente": col_name,
                                "Nom Tableau Parent": table_name,
                                "relationship": variation.get("relationship", "N/A"),
                                "isDefault": variation.get("isDefault", False),
                                "defaultHierarchy.table": variation.get("defaultHierarchy", {}).get("table", "N/A"),
                                "defaultHierarchy.hierarchy": variation.get("defaultHierarchy", {}).get("hierarchy", "N/A"),
                                "lineageTag": variation.get("lineageTag", "N/A"),
                            })

                            if "annotations" in variation:
                                for annotation in variation["annotations"]:
                                    annotations_list.append({
                                        "Type Entité Parente": "Variation Colonne",
                                        "Nom Entité Parente": variation_name,
                                        "Nom Annotation": annotation.get("name", "N/A"),
                                        "Valeur Annotation": str(annotation.get("value", "N/A")),
                                        "Nom Tableau Parent": table_name,
                                        "Nom Colonne Parent": col_name,
                                        "Nom Mesure Parent": "N/A",
                                        "Nom Relation Parente": "N/A",
                                        "Nom Hiérarchie Parente": "N/A",
                                        "Nom Niveau Hiérarchie Parente": "N/A",
                                        "Nom Partition Parente": "N/A"
                                    })

                    if "annotations" in column:
                        for annotation in column["annotations"]:
                            annotations_list.append({
                                "Type Entité Parente": "Colonne",
                                "Nom Entité Parente": col_name,
                                "Nom Annotation": annotation.get("name", "N/A"),
                                "Valeur Annotation": str(annotation.get("value", "N/A")),
                                "Nom Tableau Parent": table_name,
                                "Nom Colonne Parent": col_name,
                                "Nom Mesure Parent": "N/A",
                                "Nom Relation Parente": "N/A",
                                "Nom Hiérarchie Parente": "N/A",
                                "Nom Niveau Hiérarchie Parente": "N/A",
                                "Nom Partition Parente": "N/A"
                            })

            if "measures" in table:
                for measure in table["measures"]:
                    measure_name = measure.get("name", "Mesure sans nom")
                    measures_list.append({
                        "Nom Mesure": measure_name,
                        "Nom Tableau Parent": table_name,
                        "expression": normalize_expression(measure.get("expression", "N/A")),
                        "formatString": measure.get("formatString", "N/A"),
                        "lineageTag": measure.get("lineageTag", "N/A"),
                        "isHidden": measure.get("isHidden", False),
                        "description": measure.get("description", "N/A"),
                        "Annotations (Noms)": ", ".join([a.get("name", "Annotation sans nom") for a in measure.get("annotations", [])]),
                    })

                    if "annotations" in measure:
                        for annotation in measure["annotations"]:
                            annotations_list.append({
                                "Type Entité Parente": "Mesure",
                                "Nom Entité Parente": measure_name,
                                "Nom Annotation": annotation.get("name", "N/A"),
                                "Valeur Annotation": str(annotation.get("value", "N/A")),
                                "Nom Tableau Parent": table_name,
                                "Nom Colonne Parent": "N/A",
                                "Nom Mesure Parent": measure_name,
                                "Nom Relation Parente": "N/A",
                                "Nom Hiérarchie Parente": "N/A",
                                "Nom Niveau Hiérarchie Parente": "N/A",
                                "Nom Partition Parente": "N/A"
                            })

            if "hierarchies" in table:
                for hierarchy in table["hierarchies"]:
                    hierarchy_name = hierarchy.get("name", "Hiérarchie sans nom")
                    hierarchies_list.append({
                        "Nom Hiérarchie": hierarchy_name,
                        "Nom Tableau Parent": table_name,
                        "lineageTag": hierarchy.get("lineageTag", "N/A"),
                        "isHidden": hierarchy.get("isHidden", False),
                        "Levels (Noms)": ", ".join([l.get("name", "Niveau sans nom") for l in hierarchy.get("levels", [])]),
                        "Annotations (Noms)": ", ".join([a.get("name", "Annotation sans nom") for a in hierarchy.get("annotations", [])]),
                    })

                    if "levels" in hierarchy:
                        for level in hierarchy["levels"]:
                            level_name = level.get("name", "Niveau sans nom")
                            hierarchy_levels_list.append({
                                "Nom Niveau Hiérarchie": level_name,
                                "Nom Hiérarchie Parente": hierarchy_name,
                                "Nom Tableau Parent": table_name,
                                "ordinal": level.get("ordinal", "N/A"),
                                "column": level.get("column", "N/A"),
                                "lineageTag": level.get("lineageTag", "N/A"),
                                "isHidden": level.get("isHidden", False),
                                "Annotations (Noms)": ", ".join([a.get("name", "Annotation sans nom") for a in level.get("annotations", [])]),
                            })

                            if "annotations" in level:
                                for annotation in level["annotations"]:
                                    annotations_list.append({
                                        "Type Entité Parente": "Niveau Hiérarchie",
                                        "Nom Entité Parente": level_name,
                                        "Nom Annotation": annotation.get("name", "N/A"),
                                        "Valeur Annotation": str(annotation.get("value", "N/A")),
                                        "Nom Tableau Parent": table_name,
                                        "Nom Colonne Parente": level.get("column", "N/A"),
                                        "Nom Mesure Parente": "N/A",
                                        "Nom Relation Parente": "N/A",
                                        "Nom Hiérarchie Parente": hierarchy_name,
                                        "Nom Niveau Hiérarchie Parente": level_name,
                                        "Nom Partition Parente": "N/A"
                                    })

            if "annotations" in table:
                for annotation in table["annotations"]:
                    annotations_list.append({
                        "Type Entité Parente": "Tableau",
                        "Nom Entité Parente": table_name,
                        "Nom Annotation": annotation.get("name", "N/A"),
                        "Valeur Annotation": str(annotation.get("value", "N/A")),
                        "Nom Tableau Parent": table_name,
                        "Nom Colonne Parente": "N/A",
                        "Nom Mesure Parente": "N/A",
                        "Nom Relation Parente": "N/A",
                        "Nom Hiérarchie Parente": "N/A",
                        "Nom Niveau Hiérarchie Parente": "N/A",
                        "Nom Partition Parente": "N/A"
                    })

    if "relationships" in model_info:
        for relation in model_info["relationships"]:
            relation_name = relation.get("name", "Relation sans nom")
            relationships_list.append({
                "Nom Relation": relation_name,
                "fromTable": relation.get("fromTable", "N/A"),
                "fromColumn": relation.get("fromColumn", "N/A"),
                "toTable": relation.get("toTable", "N/A"),
                "toColumn": relation.get("toColumn", "N/A"),
                "type": relation.get("type", "N/A"),
                "crossFilteringBehavior": relation.get("crossFilteringBehavior", "N/A"),
                "isActive": relation.get("isActive", True),
                "joinOnDateBehavior": relation.get("joinOnDateBehavior", "N/A"),
                "lineageTag": relation.get("lineageTag", "N/A"),
                "Annotations (Noms)": ", ".join([a.get("name", "Annotation sans nom") for a in relation.get("annotations", [])]),
            })

            if "annotations" in relation:
                for annotation in relation["annotations"]:
                    annotations_list.append({
                        "Type Entité Parente": "Relation",
                        "Nom Entité Parente": relation_name,
                        "Nom Annotation": annotation.get("name", "N/A"),
                        "Valeur Annotation": str(annotation.get("value", "N/A")),
                        "Nom Tableau Parent": relation.get("fromTable", "N/A"),
                        "Nom Colonne Parent": "N/A",
                        "Nom Mesure Parent": "N/A",
                        "Nom Relation Parente": relation_name,
                        "Nom Hiérarchie Parente": "N/A",
                        "Nom Niveau Hiérarchie Parente": "N/A",
                        "Nom Partition Parente": "N/A"
                    })

    if "cultures" in model_info:
        for culture in model_info["cultures"]:
            culture_name = culture.get("name", "Culture sans nom")
            culture_data = {
                "Nom Culture": culture_name,
            }
            if "linguisticMetadata" in culture:
                meta = culture["linguisticMetadata"]
                if "content" in meta:
                    content = meta["content"]
                    for key, value in content.items():
                        culture_data[f"linguisticMetadata.content.{key}"] = str(value) if value is not None else "N/A"
                culture_data["linguisticMetadata.contentType"] = meta.get("contentType", "N/A")

            cultures_list.append(culture_data)

    dfs = {
        "Tables": pd.DataFrame(tables_list),
        "Partitions": pd.DataFrame(partitions_list),
        "Colonnes": pd.DataFrame(columns_list),
        "Variations Colonne": pd.DataFrame(variations_list),
        "Mesures": pd.DataFrame(measures_list),
        "Hiérarchies": pd.DataFrame(hierarchies_list),
        "Niveaux Hiérarchie": pd.DataFrame(hierarchy_levels_list),
        "Relations": pd.DataFrame(relationships_list),
        "Cultures": pd.DataFrame(cultures_list),
        "Annotations": pd.DataFrame(annotations_list),
    }

    table_order = [
        "Tables", "Partitions", "Colonnes", "Variations Colonne",
        "Mesures", "Hiérarchies", "Niveaux Hiérarchie",
        "Relations", "Cultures", "Annotations"
    ]

    ordered_dfs = collections.OrderedDict()
    for name in table_order:
        if name in dfs and not dfs[name].empty:
            ordered_dfs[name] = dfs[name]

    return ordered_dfs

def write_dfs_to_single_sheet(dfs, workbook, sheet_name="Structured Data"):
    """
    Écrit plusieurs DataFrames dans une seule feuille Excel sous forme de tableaux séparés.
    Applique le formatage et la coloration par parent.
    """
    sheet = workbook.create_sheet(sheet_name)
    current_row = 1

    global_column_widths = {}

    thin_black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    parent_colors = {}

    title_cell = sheet.cell(row=current_row, column=1, value="Modèle de Données - Vue Structurée")
    title_cell.font = Font(bold=True, size=16)
    current_row += 2

    if not dfs:
        sheet.cell(row=current_row, column=1, value="Aucune donnée structurée à afficher.").font = Font(italic=True)
        sheet.column_dimensions[get_column_letter(1)].width = (len(title_cell.value) + 4) * 1.1
        return

    for table_title, df in dfs.items():
        sheet.cell(row=current_row, column=1, value=table_title).font = Font(bold=True, size=14)
        current_row += 1

        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        if not df.empty:
            for col_index, column_name in enumerate(df.columns, start=1):
                cell = sheet.cell(row=current_row, column=col_index, value=column_name)
                cell.fill = header_fill
                cell.border = thin_black_border
                col_letter = get_column_letter(col_index)
                global_column_widths[col_letter] = max(global_column_widths.get(col_letter, 0), len(str(column_name)))
        elif len(df.columns) > 0:
            for col_index, column_name in enumerate(df.columns, start=1):
                cell = sheet.cell(row=current_row, column=col_index, value=column_name)
                cell.fill = header_fill
                cell.border = thin_black_border
                col_letter = get_column_letter(col_index)
                global_column_widths[col_letter] = max(global_column_widths.get(col_letter, 0), len(str(column_name)))

        current_row += 1

        if not df.empty:
            for row_index_in_df, row_data in df.iterrows():
                parent_identifier = None
                if "Nom Tableau Parent" in row_data and row_data["Nom Tableau Parent"] and row_data["Nom Tableau Parent"] != "N/A":
                    parent_identifier = ("Tableau", row_data["Nom Tableau Parent"])
                elif "Nom Hiérarchie Parente" in row_data and row_data["Nom Hiérarchie Parente"] and row_data["Nom Hiérarchie Parente"] != "N/A":
                    parent_identifier = ("Hiérarchie", row_data["Nom Hiérarchie Parente"])
                elif "Nom Colonne Parente" in row_data and row_data["Nom Colonne Parente"] and row_data["Nom Colonne Parente"] != "N/A":
                    parent_identifier = ("Colonne", row_data["Nom Colonne Parente"])
                elif "Nom Mesure Parent" in row_data and row_data["Nom Mesure Parent"] and row_data["Nom Mesure Parent"] != "N/A":
                    parent_identifier = ("Mesure", row_data["Nom Mesure Parent"])
                elif "Nom Partition Parente" in row_data and row_data["Nom Partition Parente"] and row_data["Nom Partition Parente"] != "N/A":
                    parent_identifier = ("Partition", row_data["Nom Partition Parente"])
                elif "Nom Relation Parente" in row_data and row_data["Nom Relation Parente"] and row_data["Nom Relation Parente"] != "N/A":
                    parent_identifier = ("Relation", row_data["Nom Relation Parente"])
                elif table_title == "Tables":
                    parent_identifier = ("Tableau", row_data["Nom Tableau"])
                elif table_title == "Relations":
                    parent_identifier = ("Relation", row_data["Nom Relation"])
                elif table_title == "Cultures":
                    parent_identifier = ("Culture", row_data["Nom Culture"])

                row_fill = None
                if parent_identifier:
                    if parent_identifier not in parent_colors:
                        parent_colors[parent_identifier] = get_distinct_color(parent_identifier[1])
                    row_fill = PatternFill(start_color=parent_colors[parent_identifier], end_color=parent_colors[parent_identifier], fill_type="solid")

                for col_index, column_name in enumerate(df.columns, start=1):
                    cell = sheet.cell(row=current_row, column=col_index, value=row_data[column_name])
                    if row_fill:
                        cell.fill = row_fill
                    cell.border = thin_black_border
                    col_letter = get_column_letter(col_index)
                    cell_value_str = str(row_data[column_name]) if row_data[column_name] is not None else ""
                    cell_length = min(len(cell_value_str), 80)
                    global_column_widths[col_letter] = max(global_column_widths.get(col_letter, 0), cell_length)

                current_row += 1

        current_row += 2

    for col_letter, width in global_column_widths.items():
        adjusted_width = (width + 2) * 0.9
        if adjusted_width > 80:
            adjusted_width = 80
        elif adjusted_width < 10:
            adjusted_width = 10
        sheet.column_dimensions[col_letter].width = adjusted_width

def run_structured_single_sheet_extraction(datamodelschema_json_path, output_directory):
    """Exécute l'extraction et le formatage des données structurées en une seule feuille."""
    excel_filename = "Data_Structure.xlsx"
    excel_output_path = os.path.join(output_directory, excel_filename)
    excel_sheet_name = "Structured Data"

    print(f"\n{'='*50}")
    print("Début de l'extraction des données structurées pour une seule feuille Excel.")
    print(f"Source JSON : {datamodelschema_json_path}")
    print(f"Fichier Excel de sortie : {excel_output_path}")

    if not datamodelschema_json_path or not os.path.exists(datamodelschema_json_path):
        print("Erreur : Le fichier DataModelSchema.json n'a pas été trouvé pour l'extraction des données structurées.")
        return False

    try:
        with open(datamodelschema_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        dfs = process_data_model_for_structured_sheet(data)

        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        write_dfs_to_single_sheet(dfs, workbook, sheet_name=excel_sheet_name)
        workbook.save(excel_output_path)

        if dfs:
            print(f"Extraction des données structurées terminée avec succès.")
            return True
        else:
            print("Aucune donnée pertinente à extraire pour les données structurées.")
            return False

    except FileNotFoundError:
        print(f"Erreur : Le fichier spécifié est introuvable à {datamodelschema_json_path} pour les données structurées.")
        return False
    except json.JSONDecodeError:
        print(f"Erreur : Impossible de décoder le fichier JSON pour les données structurées. Assurez-vous qu'il est correctement formaté.")
        return False
    except ImportError:
        print("Erreur : Les bibliothèques nécessaires (pandas ou openpyxl) ne sont pas installées.")
        print("Veuillez les installer en exécutant : pip install pandas openpyxl")
        return False
    except Exception as e:
        print(f"Une erreur est survenue lors de l'extraction des données structurées : {e}")
        return False

# --- Fonctions pour l'extraction des KPIs ---

def find_executable(executable_name):
    """Recherche un exécutable uniquement dans Downloads, Desktop et le dossier cible."""
    search_paths = [
        os.path.expanduser("~/Downloads"),           # Dossier Downloads
        os.path.expanduser("~/Desktop"),            # Bureau (Desktop)
        output_directory                            # Data Extractor (dossier cible)
    ]

    for path in search_paths:
        if os.path.exists(path):
            for root, _, files in os.walk(path):
                if executable_name in files:
                    full_path = os.path.join(root, executable_name)
                    print(f"Trouvé '{executable_name}' dans : {root}")
                    return full_path

    # Si non trouvé, afficher un popup
    root = tk.Tk()
    root.withdraw()  # Masquer la fenêtre principale
    messagebox.showerror("Erreur", "Téléchargez et Installez pbi-tools.exe et pbi-tools.core.exe et déplacez les dans Data_Extractor.")
    root.destroy()
    return None

def find_layout_file(directory):
    """Recherche le fichier 'Layout' (sans extension) dans un répertoire et ses sous-répertoires."""
    for root, _, files in os.walk(directory):
        if 'Layout' in files:
            return os.path.join(root, 'Layout')
    return None

def find_datamodelschema_file(directory):
    """Recherche le fichier 'DataModelSchema' dans un répertoire et ses sous-répertoires."""
    for root, _, files in os.walk(directory):
        if 'DataModelSchema' in files:
            return os.path.join(root, 'DataModelSchema')
    return None

def extract_layout_json_from_pbix_or_file(source_file_path, output_dir):
    """Extrait le fichier 'Layout' d'un fichier Power BI (.pbix ou .file) comme Layout.json."""
    print(f"Extraction du fichier Layout.json à partir du fichier Power BI.")
    if not os.path.exists(source_file_path):
        print(f"Erreur : Le fichier source n'existe pas : {source_file_path}")
        return None

    os.makedirs(output_dir, exist_ok=True)
    json_files_dir = os.path.join(output_dir, "JSON Files")
    os.makedirs(json_files_dir, exist_ok=True)
    temp_dir = None
    extracted_layout_path = None
    layout_output_path = os.path.join(json_files_dir, 'Layout.json')

    try:
        temp_dir = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(source_file_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
        except zipfile.BadZipFile:
            print(f"Erreur : Le fichier '{os.path.basename(source_file_path)}' ne semble pas être une archive ZIP valide.")
            return None
        except Exception as e:
            print(f"Erreur lors de l'ouverture ou de l'extraction de l'archive : {e}")
            return None

        extracted_layout_path_temp = find_layout_file(temp_dir)
        if not extracted_layout_path_temp:
            print(f"Erreur : Le fichier 'Layout' n'a pas été trouvé dans l'archive ou ses sous-dossiers.")
            return None

        content = read_file_with_multiple_encodings(extracted_layout_path_temp)
        if content:
            try:
                start_idx = content.find('{')
                if start_idx > 0:
                    content = content[start_idx:]
                data = json.loads(content)
                with open(layout_output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                print(f"Fichier 'Layout.json' extrait avec succès.")
                extracted_layout_path = layout_output_path
            except json.JSONDecodeError as e:
                print(f"Erreur : Le contenu du fichier 'Layout' n'est pas un JSON valide : {e}")
                return None
            except Exception as e:
                print(f"Erreur lors de la sauvegarde du fichier JSON : {e}")
                return None
        else:
            print(f"Erreur : Impossible de lire le contenu du fichier 'Layout'.")
            return None

    finally:
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

    return extracted_layout_path

def extract_datamodelschema_from_pbix(source_file_path, output_dir, pbi_tools_path, pbi_tools_core_path):
    """
    Convertit un fichier .pbix en .pbit à l'aide de pbi-tools, extrait DataModelSchema
    du .pbit, et le sauvegarde en tant que DataModelSchema.json dans le répertoire de sortie.
    """
    print("Extraction du fichier DataModelSchema.json à partir du fichier Power BI.")
    try:
        if not os.path.exists(source_file_path):
            print(f"Erreur : Le fichier .pbix n'a pas été trouvé dans : {source_file_path}")
            return None

        if not os.path.exists(pbi_tools_path):
            print(f"Erreur : L'exécutable pbi-tools n'a pas été trouvé dans le chemin spécifié : {pbi_tools_path}")
            return None

        if not os.path.exists(pbi_tools_core_path):
            print(f"Erreur : L'exécutable pbi-tools.core n'a pas été trouvé dans : {pbi_tools_core_path}")
            return None

        os.makedirs(output_dir, exist_ok=True)
        json_files_dir = os.path.join(output_dir, "JSON Files")
        os.makedirs(json_files_dir, exist_ok=True)
        temp_dir = None
        default_extract_folder = os.path.splitext(source_file_path)[0]
        output_pbit_path = None
        extracted_datamodelschema_path = None
        datamodelschema_output_path = os.path.join(json_files_dir, 'DataModelSchema.json')

        try:
            temp_dir = tempfile.mkdtemp()

            if os.path.exists(default_extract_folder):
                print(f"Suppression du dossier d'extraction existant créé par pbi-tools : {os.path.basename(default_extract_folder)}")
                shutil.rmtree(default_extract_folder, ignore_errors=True)

            print("Exécution de pbi-tools extract pour générer les données brutes.")
            try:
                cmd = [pbi_tools_path, "extract", source_file_path, "-modelSerialization", "Raw"]
                result = subprocess.run(cmd, capture_output=True, text=True, check=True, creationflags=subprocess.CREATE_NO_WINDOW)
                if result.stderr:
                    print(f"Erreurs stderr : {result.stderr.strip()}")
                if os.path.exists(default_extract_folder):
                    print(f"Extraction pbi-tools réussie dans : {os.path.basename(default_extract_folder)}/")
                else:
                    print(f"Erreur : Dossier d'extraction par défaut non trouvé après l'exécution.")
                    return None
            except subprocess.CalledProcessError as e:
                print(f"Erreur lors de l'extraction de pbi-tools (code {e.returncode}) :")
                print(f"Sortie : {e.stdout.strip()}")
                print(f"Erreur : {e.stderr.strip()}")
                return None
            except Exception as e:
                print(f"Erreur inattendue lors de l'exécution de pbi-tools extract : {e}")
                return None

            datamodelschema_path = find_datamodelschema_file(default_extract_folder)
            if datamodelschema_path:
                print(f"Fichier DataModelSchema trouvé directement dans l'extraction Raw.")
                content = read_file_with_multiple_encodings(datamodelschema_path)
                if content:
                    try:
                        data = json.loads(content)
                        with open(datamodelschema_output_path, 'w', encoding='utf-8') as f:
                            json.dump(data, f, indent=4, ensure_ascii=False)
                        print(f"Fichier DataModelSchema.json extrait avec succès.")
                        extracted_datamodelschema_path = datamodelschema_output_path
                        return extracted_datamodelschema_path
                    except json.JSONDecodeError as e:
                        print(f"Erreur : Le contenu de DataModelSchema n'est pas un JSON valide : {e}")
                        return None
                    except Exception as e:
                        print(f"Erreur lors de la sauvegarde de DataModelSchema.json : {e}")
                        return None
                else:
                    print(f"Erreur : Impossible de lire DataModelSchema avec les encodages essayés.")
                    return None
            else:
                print(f"Aucun fichier DataModelSchema trouvé dans l'extraction Raw. Tentative de compilation en .pbit pour l'extraire.")

            model_folder = os.path.join(default_extract_folder, 'Model')
            if not os.path.exists(model_folder) or not os.listdir(model_folder):
                print(f"Avertissement : Aucun dossier 'Model' ou modèle de données trouvé dans l'extraction. Le fichier .pbix peut ne pas contenir de modèle de données à compiler.")
                return None

            print("Exécution de pbi-tools.core compile pour générer un fichier .pbit.")
            output_pbit_path = os.path.join(temp_dir, os.path.splitext(os.path.basename(source_file_path))[0] + '.pbit')
            try:
                cmd = [pbi_tools_core_path, "compile", default_extract_folder, output_pbit_path, "PBIT", "True"]
                result = subprocess.run(cmd, capture_output=True, text=True, check=True, creationflags=subprocess.CREATE_NO_WINDOW)
                if result.stderr:
                    print(f"Erreurs stderr : {result.stderr.strip()}")
                print(f"Compilation réussie en {os.path.basename(output_pbit_path)}.")
            except subprocess.CalledProcessError as e:
                print(f"Erreur lors de l'exécution de pbi-tools.core compile (code {e.returncode}) :")
                print(f"Sortie : {e.stdout.strip()}")
                print(f"Erreur : {e.stderr.strip()}")
                return None
            except Exception as e:
                print(f"Erreur inattendue lors de l'exécution de pbi-tools.core compile : {e}")
                return None

            if not os.path.exists(output_pbit_path):
                print(f"Erreur : Le fichier .pbit n'a pas été créé à l'emplacement : {output_pbit_path}")
                return None

            print("Extraction de DataModelSchema depuis le fichier .pbit généré.")
            try:
                with zipfile.ZipFile(output_pbit_path, 'r') as zip_ref:
                    datamodelschema_path_in_zip = None
                    for file_name in zip_ref.namelist():
                        if file_name == 'DataModelSchema':
                            datamodelschema_path_in_zip = file_name
                            break
                    if not datamodelschema_path_in_zip:
                        print(f"Erreur : Fichier DataModelSchema non trouvé dans l'archive .pbit.")
                        return None

                    zip_ref.extract(datamodelschema_path_in_zip, temp_dir)
                    extracted_file_path = os.path.join(temp_dir, datamodelschema_path_in_zip)
                    content = read_file_with_multiple_encodings(extracted_file_path)
                    if content:
                        try:
                            data = json.loads(content)
                            with open(datamodelschema_output_path, 'w', encoding='utf-8') as f:
                                json.dump(data, f, indent=4, ensure_ascii=False)
                            print(f"Fichier DataModelSchema.json extrait avec succès.")
                            extracted_datamodelschema_path = datamodelschema_output_path
                        except json.JSONDecodeError as e:
                            print(f"Erreur : Le contenu de DataModelSchema n'est pas un JSON valide : {e}")
                            return None
                        except Exception as e:
                            print(f"Erreur lors de la sauvegarde de DataModelSchema.json : {e}")
                            return None
                    else:
                        print(f"Erreur : Impossible de lire DataModelSchema avec les encodages essayés.")
                        return None
            except zipfile.BadZipFile:
                print(f"Erreur : Le fichier .pbit '{os.path.basename(output_pbit_path)}' n'est pas une archive ZIP valide.")
                return None
            except Exception as e:
                print(f"Erreur lors de l'extraction de l'archive .pbit : {e}")
                return None

        finally:
            if os.path.exists(default_extract_folder):
                print(f"Nettoyage du dossier d'extraction : {os.path.basename(default_extract_folder)}")
                shutil.rmtree(default_extract_folder, ignore_errors=True)

            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)

    except Exception as e:
        print(f"Erreur générale dans extract_datamodelschema_from_pbix : {e}")
        return None

    return extracted_datamodelschema_path

def extract_table_from_queryref(query_ref):
    """Extrait le nom de la table du queryRef."""
    if not isinstance(query_ref, str):
        return "N/A"

    match = re.search(r'\(([^.]+)\.|^([^.]+)\.', query_ref)
    if match:
        table_name = match.group(1) if match.group(1) else match.group(2)
        return table_name.strip()
    return "N/A"

def extract_all_kpis_from_powerbi_report(json_file_path):
    """Extrait les KPIs (mesures calculées) des données JSON de Layout."""
    print("Analyse du fichier Layout.json pour extraire les KPIs.")
    try:
        with open(json_file_path, 'r', encoding='utf-16') as file:
            content = file.read()
            start_idx = content.find('{')
            if start_idx > 0:
                content = content[start_idx:]
            data = json.loads(content)
    except UnicodeDecodeError:
        print("Erreur de décodage Unicode. Tentative avec l'encodage 'utf-8'.")
        try:
            with open(json_file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                start_idx = content.find('{')
                if start_idx > 0:
                    content = content[start_idx:]
                data = json.loads(content)
        except json.JSONDecodeError as e:
            print(f"Erreur de décodage JSON : {e}")
            return None
        except Exception as e:
            print(f"Erreur lors de la lecture du fichier avec utf-8 : {e}")
            return None
    except json.JSONDecodeError as e:
        print(f"Erreur de décodage JSON : {e}")
        return None
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier avec utf-16 : {e}")
        return None

    all_kpis = []
    sections = data.get('sections', [])
    for section_index, section in enumerate(sections):
        section_name = section.get('displayName', f'Section {section_index + 1}')
        for visual_container in section.get('visualContainers', []):
            config_str = visual_container.get('config', '{}')
            try:
                config = json.loads(config_str)
                visual_type = None
                if 'singleVisual' in config and 'visualType' in config['singleVisual']:
                    visual_type = config['singleVisual']['visualType']
                if 'singleVisual' in config and 'projections' in config['singleVisual']:
                    projections = config['singleVisual']['projections']
                    for role, items in projections.items():
                        for item in items:
                            if 'queryRef' in item:
                                query_ref = item['queryRef']
                                base_name = query_ref
                                alias = ""
                                dax_formula = query_ref
                                is_calculated = False
                                source_table = extract_table_from_queryref(query_ref)
                                data_transforms_str = visual_container.get('dataTransforms', '{}')
                                try:
                                    data_transforms = json.loads(data_transforms_str)
                                    if 'selects' in data_transforms:
                                        for select in data_transforms['selects']:
                                            if select.get('queryName') == query_ref:
                                                alias = select.get('displayName', "")
                                                if 'expr' in select:
                                                    expr = select['expr']
                                                    if 'Aggregation' in str(expr) or 'Measure' in str(expr):
                                                        is_calculated = True
                                                    elif 'Column' in str(expr) or 'HierarchyLevel' in str(expr):
                                                        is_calculated = False
                                except json.JSONDecodeError:
                                    pass
                                measure_type = "Mesure Calculée" if is_calculated else "Mesure non Calculée"
                                kpi = {
                                    "Nom de Base": base_name,
                                    "Alias Power BI": alias,
                                    "Formule DAX": dax_formula,
                                    "Type Visuel": visual_type,
                                    "Type Mesure": measure_type,
                                    "Source Table": source_table,
                                    "Source": f"Visuel ({section_name})",
                                }
                                all_kpis.append(kpi)
            except json.JSONDecodeError:
                continue

    model_kpis = []
    def find_measures_in_json(data):
        measures = []
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'measures' and isinstance(value, list):
                    for measure_def in value:
                        name = measure_def.get('name', 'N/A')
                        expression = measure_def.get('expression', 'N/A')
                        display_name = measure_def.get('properties', {}).get('dataViewDisplayName', name)
                        source_table = "N/A (Modèle)"
                        measures.append({
                            "Nom de Base": name,
                            "Alias Power BI": display_name,
                            "Formule DAX": expression,
                            "Type Visuel": "N/A",
                            "Type Mesure": "Mesure Calculée",
                            "Source Table": source_table,
                            "Source": "Modèle (potentiel)",
                        })
                elif isinstance(value, (dict, list)):
                    measures.extend(find_measures_in_json(value))
        elif isinstance(data, list):
            for item in data:
                measures.extend(find_measures_in_json(item))
        return measures

    model_kpis = find_measures_in_json(data)
    final_kpis_list = all_kpis.copy()
    for model_kpi in model_kpis:
        found_in_visuals = False
        for visual_kpi in final_kpis_list:
            if visual_kpi['Nom de Base'] == model_kpi['Nom de Base']:
                visual_kpi['Source'] += " et Modèle (potentiel)"
                if model_kpi['Formule DAX'] != 'N/A' and visual_kpi['Formule DAX'] == visual_kpi['Nom de Base']:
                    visual_kpi['Formule DAX'] = model_kpi['Formule DAX']
                found_in_visuals = True
                break
        if not found_in_visuals:
            final_kpis_list.append(model_kpi)

    kpis_df = pd.DataFrame(final_kpis_list)
    if 'Type Mesure' in kpis_df.columns:
        kpis_df = kpis_df[kpis_df['Type Mesure'] == 'Mesure Calculée'].copy()
        kpis_df = kpis_df.drop('Type Mesure', axis=1)

    cols = ["Nom de Base", "Alias Power BI", "Source Table", "Formule DAX", "Type Visuel", "Source"]
    cols = [col for col in cols if col in kpis_df.columns]
    kpis_df = kpis_df[cols]
    if kpis_df is not None and not kpis_df.empty:
        print("Extraction des KPIs terminée avec succès.")
    else:
        print("Aucun KPI pertinent (Mesure Calculée) n'a été extrait.")
    return kpis_df

def run_kpi_extraction(output_directory):
    """
    Exécute l'extraction des KPIs depuis un fichier PBIX et retourne le DataFrame des KPIs.
    Retourne également le chemin du DataModelSchema.json extrait si succès.
    """
    print(f"\n{'='*50}")
    print("Début de l'extraction des KPIs à partir d'un fichier Power BI.")
    
    layout_json_path_in_output = os.path.join(output_directory, "JSON Files", "Layout.json")
    datamodelschema_json_path_in_output = os.path.join(output_directory, "JSON Files", "DataModelSchema.json")

    extracted_datamodelschema_file_path = None

    try:
        print("Recherche des exécutables pbi-tools pour l'extraction des KPIs.")
        pbi_tools_path = find_executable("pbi-tools.exe")
        pbi_tools_core_path = find_executable("pbi-tools.core.exe")

        if not pbi_tools_path or not pbi_tools_core_path:
            raise FileNotFoundError("Les exécutables pbi-tools.exe et/ou pbi-tools.core.exe n'ont pas été trouvés.")

        print("Exécutables pbi-tools trouvés et prêts à être utilisés.")
        # Déplacer les exécutables vers un sous-dossier pbi-tools dans Data Extractor
        pbi_tools_subdir = os.path.join(output_directory, "pbi-tools")
        os.makedirs(pbi_tools_subdir, exist_ok=True)
        try:
            shutil.move(pbi_tools_path, os.path.join(pbi_tools_subdir, "pbi-tools.exe"))
            shutil.move(pbi_tools_core_path, os.path.join(pbi_tools_subdir, "pbi-tools.core.exe"))
            pbi_tools_path = os.path.join(pbi_tools_subdir, "pbi-tools.exe")
            pbi_tools_core_path = os.path.join(pbi_tools_subdir, "pbi-tools.core.exe")
            print(f"Déplacement des exécutables vers {pbi_tools_subdir} effectué avec succès.")
        except Exception as e:
            print(f"Erreur lors du déplacement des exécutables : {e}. Utilisation des chemins originaux.")

        print("Ouverture de la fenêtre pour sélectionner un fichier Power BI.")
        root = tk.Tk()
        root.withdraw()
        filetypes = [("Power BI Files", "*.pbix *.file"), ("All files", "*.*")]
        source_powerbi_file = filedialog.askopenfilename(
            title="Sélectionnez le fichier Power BI (.pbix or .file) pour l'extraction des KPIs et du schéma de données",
            filetypes=filetypes
        )
        root.destroy()

        if not source_powerbi_file:
            print("Aucun fichier Power BI sélectionné. Extraction des KPIs et du schéma annulée.")
            return None, None

        print(f"Fichier Power BI sélectionné : {os.path.basename(source_powerbi_file)}.")

        print("Extraction du fichier Layout.json à partir du fichier Power BI.")
        extracted_layout_file = extract_layout_json_from_pbix_or_file(source_powerbi_file, output_directory)
        print("Extraction du fichier DataModelSchema.json à partir du fichier Power BI.")
        extracted_datamodelschema_file = extract_datamodelschema_from_pbix(
            source_file_path=source_powerbi_file,
            output_dir=output_directory,
            pbi_tools_path=pbi_tools_path,
            pbi_tools_core_path=pbi_tools_core_path
        )

        extracted_datamodelschema_file_path = extracted_datamodelschema_file

        if extracted_layout_file and os.path.exists(extracted_layout_file):
            df_kpis = extract_all_kpis_from_powerbi_report(extracted_layout_file)
            if df_kpis is not None and not df_kpis.empty:
                return df_kpis, extracted_datamodelschema_file_path
            else:
                return None, extracted_datamodelschema_file_path
        else:
            print("Échec de l'extraction de Layout.json. Extraction des KPIs annulée.")
            return None, extracted_datamodelschema_file_path

    except FileNotFoundError as e:
        print(f"Erreur : Un fichier exécutable pbi-tools nécessaire n'a pas été trouvé : {e}")
        print("Veuillez placer 'pbi-tools.exe' et 'pbi-tools.core.exe' dans Downloads, le répertoire du script, ou Data Extractor.")
        return None, None
    except Exception as e:
        print(f"Une erreur est survenue lors de l'extraction des KPIs : {e}")
        return None, None

# --- Fonction pour fusionner les DataFrames dans Extracted_Data.xlsx ---

def merge_excel_files(df_tables, df_kpis, output_directory):
    """
    Fusionne les DataFrames des tables/colonnes et des KPIs dans un fichier Excel unique
    avec des onglets séparés ('Données Granulaires' et 'KPIs').
    """
    print(f"\n{'='*50}")
    print("Début de la fusion des données dans un fichier Excel unique.")
    
    output_file = os.path.join(output_directory, "Extracted_Data.xlsx")
    try:
        wb = Workbook()
        # Style pour l'en-tête (jaune, gras, bordures)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin', color='000000'), 
                             right=Side(style='thin', color='000000'), 
                             top=Side(style='thin', color='000000'), 
                             bottom=Side(style='thin', color='000000'))
        # Bordure standard pour les cellules de données
        data_border = Border(left=Side(style='thin', color='000000'), 
                            right=Side(style='thin', color='000000'), 
                            top=Side(style='thin', color='000000'), 
                            bottom=Side(style='thin', color='000000'))

        # Écriture du DataFrame des tables/colonnes
        if df_tables is not None and not df_tables.empty:
            print("Écriture des données granulaires dans l'onglet 'Données Granulaires'.")
            ws_tables = wb.create_sheet("Données Granulaires", 0)
            # Écrire l'en-tête
            for col_idx, col_name in enumerate(df_tables.columns, 1):
                cell = ws_tables.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            # Écrire les données avec la couleur basée sur 'Nom de la Table'
            unique_tables = df_tables['Nom de la Table'].unique()
            table_colors = {table: get_distinct_color(table) for table in unique_tables}
            for row_idx, row in df_tables.iterrows():
                table_name = row['Nom de la Table']
                fill_color = PatternFill(start_color=table_colors[table_name], end_color=table_colors[table_name], fill_type="solid")
                for col_idx, value in enumerate(row, 1):
                    cell = ws_tables.cell(row=row_idx + 2, column=col_idx)
                    cell.value = value
                    cell.fill = fill_color
                    cell.border = data_border
            # Ajuster la largeur des colonnes
            for col in ws_tables.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 4) * 1.1, 100)
                ws_tables.column_dimensions[column].width = max(adjusted_width, 10)
        else:
            print("Aucune donnée de tables/colonnes à écrire dans l'onglet 'Données Granulaires'.")

        # Écriture du DataFrame des KPIs (corrigé pour éviter les lignes vides)
        if df_kpis is not None and not df_kpis.empty:
            print("Écriture des KPIs dans l'onglet 'KPIs'.")
            ws_kpis = wb.create_sheet("KPIs", 1)
            # Définir les couleurs pour la colonne 'Source'
            colors = [
                'D9E1F2', 'E2EFDA', 'FFF2CC', 'FCE4D6',
                'E7E6E6', 'FBE4D5', 'C6E0B4', 'BDD7EE'
            ]
            if 'Source' in df_kpis.columns:
                unique_sources = df_kpis['Source'].unique()
                source_colors = {source: colors[i % len(colors)] for i, source in enumerate(unique_sources)}
            else:
                source_colors = {}
                print("Colonne 'Source' non trouvée dans les données KPIs. Utilisation de la couleur par défaut.")
            # Écrire l'en-tête
            for col_idx, col_name in enumerate(df_kpis.columns, 1):
                cell = ws_kpis.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            # Écrire les données sans lignes vides
            for row_idx in range(len(df_kpis)):
                row_data = df_kpis.iloc[row_idx]
                source_value = row_data['Source'] if 'Source' in df_kpis.columns else 'Default'
                fill_color = PatternFill(start_color=source_colors.get(source_value, 'FFFFFF'), 
                                        end_color=source_colors.get(source_value, 'FFFFFF'), 
                                        fill_type="solid")
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_kpis.cell(row=row_idx + 2, column=col_idx)
                    cell.value = value if pd.notna(value) else ""
                    cell.fill = fill_color
                    cell.border = data_border
            # Ajuster la largeur des colonnes
            for col in ws_kpis.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 4) * 1.1, 100)
                ws_kpis.column_dimensions[column].width = max(adjusted_width, 10)
        else:
            print("Aucune donnée de KPIs à écrire dans l'onglet 'KPIs'.")

        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Sauvegarder le fichier fusionné
        wb.save(output_file)
        print(f"Fichier Excel 'Extracted_Data.xlsx' généré avec succès.")
        return True

    except Exception as e:
        print(f"Erreur lors de la fusion des données dans 'Extracted_Data.xlsx' : {str(e)}")
        if os.path.exists(output_file):
            try:
                os.remove(output_file)
                print(f"Fichier corrompu '{os.path.basename(output_file)}' supprimé.")
            except:
                print(f"Impossible de supprimer le fichier corrompu '{os.path.basename(output_file)}'.")
        return False

# --- Point d'entrée principal ---

if __name__ == "__main__":
    print("Début de l'exécution du script d'extraction de données Power BI.")
    print(f"Répertoire de sortie configuré : {output_directory}")

    # Nettoyage des fichiers existants pour éviter les confusions
    data_structure_path = os.path.join(output_directory, "Data_Structure.xlsx")
    extracted_data_path = os.path.join(output_directory, "Extracted_Data.xlsx")
    layout_json_path = os.path.join(output_directory, "JSON Files", "Layout.json")
    datamodelschema_json_path = os.path.join(output_directory, "JSON Files", "DataModelSchema.json")

    if os.path.exists(data_structure_path):
        os.remove(data_structure_path)
        print(f"Fichier précédent 'Data_Structure.xlsx' supprimé.")
    if os.path.exists(extracted_data_path):
        os.remove(extracted_data_path)
        print(f"Fichier précédent 'Extracted_Data.xlsx' supprimé.")
    if os.path.exists(layout_json_path):
        os.remove(layout_json_path)
        print(f"Fichier précédent 'Layout.json' supprimé.")
    if os.path.exists(datamodelschema_json_path):
        os.remove(datamodelschema_json_path)
        print(f"Fichier précédent 'DataModelSchema.json' supprimé.")

    try:
        print("Lancement de l'extraction des KPIs à partir du fichier Power BI.")
        df_kpis, extracted_datamodelschema_path = run_kpi_extraction(output_directory)

        extraction_success = False
        structured_success = False
        merge_success = False

        if extracted_datamodelschema_path and os.path.exists(extracted_datamodelschema_path):
            print("Lancement de l'extraction des tables et colonnes visibles.")
            df_tables = run_tables_columns_extraction(extracted_datamodelschema_path, output_directory)

            print("Lancement de l'extraction des données structurées pour une seule feuille.")
            structured_success = run_structured_single_sheet_extraction(extracted_datamodelschema_path, output_directory)

            print("Lancement de la fusion des données dans un fichier Excel unique.")
            if df_tables is not None or df_kpis is not None:
                merge_success = merge_excel_files(df_tables, df_kpis, output_directory)
            else:
                print("Aucune donnée extraite pour générer le fichier Excel.")
        else:
            print("Le fichier DataModelSchema.json n'a pas été extrait avec succès.")
            print("Les extractions des tables/colonnes et des données structurées ne peuvent pas être effectuées.")
            print("La fusion des données ne peut pas être réalisée.")
            extraction_success = False
        extraction_success = (extracted_datamodelschema_path is not None and os.path.exists(extracted_datamodelschema_path))

    except Exception as e:
        print(f"Erreur lors de l'exécution du script : {e}")
        import traceback
        traceback.print_exc()

    # Affichage de la popup avec icônes
    root = tk.Tk()
    root.withdraw()

    # Obtenir la date et l'heure actuelles
    current_time = datetime.now().strftime("%H:%M %z, %d/%m/%Y")
    message = f"Résumé de l'exécution ({current_time}) : \n\n"
    success_icon = "✔"
    failure_icon = "✘"

    # Vérification des résultats
    if extraction_success and structured_success and merge_success:
        message += f"- Extraction de la structure de données (Data_Structure) : {success_icon} [Succès]\n"
        message += f"- Extraction des données granulaire + KPIs (Extracted_Data.xlsx) : {success_icon} [Succès]\n"
    elif extraction_success and structured_success and not merge_success:
        message += f"- Extraction de la structure de données (Data_Structure) : {success_icon} [Succès]\n"
        message += f"- Extraction des données granulaire + KPIs (Extracted_Data.xlsx) : {failure_icon} [Échec]\n"
    elif extraction_success and not structured_success and not merge_success:
        message += f"- Extraction de la structure de données (Data_Structure) : {failure_icon} [Échec]\n"
        message += f"- Extraction des données granulaire + KPIs (Extracted_Data.xlsx) : {failure_icon} [Échec]\n"
    elif not extraction_success:
        message += f"- Extraction de la structure de données (Data_Structure) : {failure_icon} [Échec]\n"
        message += f"- Extraction des données granulaire + KPIs (Extracted_Data.xlsx) : {failure_icon} [Échec]\n"

    # Vérification des fichiers JSON uniquement s'ils n'ont pas été extraits
    if not os.path.exists(layout_json_path) and not os.path.exists(datamodelschema_json_path):
        message += f"- Extraction des fichiers JSON (Layout.json et DataModelSchema.json) : {failure_icon} [Échec]\n"
    elif not os.path.exists(layout_json_path) and os.path.exists(datamodelschema_json_path):
        message += f"- Extraction du fichier JSON 'Layout.json' : {failure_icon} [Échec]\n"
    elif os.path.exists(layout_json_path) and not os.path.exists(datamodelschema_json_path):
        message += f"- Extraction du fichier JSON 'DataModelSchema.json' : {failure_icon} [Échec]\n"

    messagebox.showinfo("Résultat de l'Extraction", message, parent=root)
    root.destroy()

    print("Exécution du script d'extraction de données Power BI terminée.")